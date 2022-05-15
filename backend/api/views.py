from django.http import HttpResponse
from rest_framework.generics import ListAPIView
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework.authentication import BasicAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions
from rest_framework.pagination import PageNumberPagination

from main.models import *

from django.db import transaction

from .serializer import *

import traceback

import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from openpyxl.writer.excel import save_virtual_workbook
from django.conf import settings
import os

from django.contrib.auth.decorators import permission_required

def get_user_role(request):
    # roles by groups: texniklar=0, tuman_administratorlari = 1 for just now
    group = request.user.groups.all()
    role = 0
    if len(group) > 0:
        role_name = group[0].name

        if role_name == "tuman_administratorlari":
            role = 1
    
    return role

def get_user_permissions(user):
    permissions = []
    for group in Group.objects.filter(user=user):
        for per in group.permissions.all():
            permissions.append(per.codename)
    return permissions

class CountryListView(ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

class RegionListView(ListAPIView):
    serializer_class = CountrySerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        """
        This view should return a list of all the regions
        for the currently country.
        """
        countryId = self.kwargs['countryId']
        return Region.objects.filter(country_id=countryId)

class DistrictListView(ListAPIView):
    serializer_class = DistrictSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        """
        This view should return a list of all the districts
        for the currently region.
        """
        user_districts = self.request.user.staff.districts.all()
        return user_districts

class FarmListView(ListAPIView):
    serializer_class = FarmSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        """
        This view should return a list of all the districts
        for the currently region.
        """
        
        id = int(self.kwargs['id'])
        if id != 0:
            return Farm.objects.filter(pk = id)
        districts = self.request.user.staff.districts.all()
        if len(districts) == 0:
            return Farm.objects.all()
        q = Farm.objects.filter(district=districts[0])
        for i in range(1, len(districts)):
            q += Farm.objects.filter(district=districts[i])
        return q
        
class WellListView(ListAPIView):
    serializer_class = WellSerializer
    permission_classes = [DjangoModelPermissions]
    def get_queryset(self):
        id = int(self.kwargs['id'])
        
        role = get_user_role(self.request)
        # Respond all well which are connected to texniklar
        if role == 0:
            return Well.objects.filter(user = self.request.user)
        elif role == 1:
            farms = Farm.objects.filter(district = self.request.user.staff.district)

            query = Well.objects.filter(farm = farms[0])
            for i in range(1, len(farms)):
                query = query.union(Well.objects.filter(farm = farms[i]))
            return query
    
        return Well.objects.all()[:10]

class MusterPumpingListView(ListAPIView):
    serializer_class = MusterPumpingSerializer
    permission_classes = [DjangoModelPermissions]
    def get_queryset(self):
        if "well_id" in self.kwargs:
            well_id = int(self.kwargs['well_id'])
            return MusterPumping.objects.filter(well=well_id)
        return MusterPumping.objects.filter()

#class YieldSizeViewSet

@api_view(['GET'])
def farm_request(request, district_id=0):
    if request.method == "GET":
        return Response(FarmSerializer(Farm.objects.filter(district=district_id), many=True).data, status=200)

@api_view(['POST', 'DELETE', 'PUT'])
def farm_request1(request):
    
    if request.method == "POST":
        ser_obj = FarmSerializer(data=request.data)
        if ser_obj.is_valid():
            ser_obj.save()
            return Response(ser_obj.data.get('id'), status=201)
        else:
            return Response("Not acceptable", status=406)
    print(request.data)

    # find suitable instance
    try:
        instance = Farm.objects.get(id=request.data.get('id'))
    except:
        return Response(status=404)
     # Update object
    if request.method == 'PUT':
        serializer = FarmSerializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=200)
        else:
            return Response(serializer.errors, status=406)
    else:
        instance.delete()
        return Response('Deleted', status=200)

@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def well_request(request, id=0):
    try:
        if request.method == 'GET':
            if id != 0:
                well = Well.objects.get(pk = id)
                return Response(WellSerializer(well).data, status=200)
            
            # if the user is technician, then must send the well which were attached to the user
            if request.user.has_perm('main.add_musterpumping'):
                query = Well.objects.filter(user=request.user)
                return Response(WellSerializer(query, many=True).data, status=200)
            
            # if the user is the district controller, then must send the all well in the district
            if request.user.has_perm('main.add_well'):
                # the user is the district controller, then the user must have one district, so we get the first one
                farms = Farm.objects.filter(district = request.user.staff.districts.all()[0])
                query = Well.objects.filter(farm = farms[0])

                for i in range(1, len(farms)):
                    query = query.union(Well.objects.filter(farm = farms[i]))
                return Response(WellSerializer(query, many=True).data, status=200)
            return Response("Not have permission", status=403)
        # Create object
        if request.method == 'POST':
            request.data['user']=request.user.id 
            serializer = WellSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)
        
        # find suitable instance
        try:
            instance = Well.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            request.data['user']=request.user.id 
            serializer = WellSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(['POST', 'DELETE', 'PUT'])
def muster_pumping_request(request):
    try:
        if request.method == "POST":
            permission = request.user.has_perm('main.add_musterpumping')
            if permission == False:
                return Response("You do not have permission", status=403)
            
            serializer = MusterPumpingSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        try:
            instance = MusterPumping.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            permission = request.user.has_perm('main.change_musterpumping')
            if permission == False:
                return Response("You do not have permission", status=403)

            request.data['user']=request.user.id 
            serializer = MusterPumpingSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            permission = request.user.has_perm('main.delete_musterpumping')
            if permission == False:
                return Response("You do not have permission", status=403)
            instance.delete()
            return Response('Deleted', status=200)
    except:
        return Response(status=500)

@api_view(['GET'])
def district_request(request):
    if request.method == "GET":
        return Response(DistrictSerializer(District.objects.all(), many=True).data, status=200)

@api_view(['GET'])
def get_user_district(request):
    if request.method == "GET":
        return Response(DistrictSerializer(request.user.staff.districts, many=True).data, status=200)

@api_view(["GET"])
def user_permissions(request, app_name, model_name):
    permissions = {
        'view': request.user.has_perm(app_name + '.view_' + model_name),
        'add': request.user.has_perm(app_name + '.add_' + model_name),
        'change': request.user.has_perm(app_name + '.change_' + model_name),
        'delete': request.user.has_perm(app_name + '.delete_' + model_name),
    }
    return Response(permissions, status=200)

@api_view(["GET"])
def user_groups(request):
    groups = []
    for item in request.user.groups.all():
        groups.append(item.name)
    return Response(groups, status=200)

@api_view(["GET"])
def user_role(request):
    role = get_user_role(request)
    return Response(role, status=200)

@api_view(["GET"])
def get_technicians(request):
    try:
        current_staff = Staff.objects.get(user=request.user)
        districts = current_staff.districts.all()
        if len(districts) > 0:
            technicians = Staff.objects.filter(districts=districts[0]).exclude(id=current_staff.id)
            for i in range(1, len(districts)):
                technicians += Staff.objects.filter(districts=districts[i])
            technicians = technicians.filter(position__priority__lt = current_staff.position.priority)
        return Response(StaffSerializer(technicians, many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response("Not found", status=404)

@api_view(["PUT"])
def attach_well_to_technician(request):
    wells = request.data['wells']
    technician = request.data['technician']
    technician = Staff.objects.get(pk=technician)
    
    with transaction.atomic():
        for well_id in wells:
            well = Well.objects.get(pk=well_id)
            well.user = technician.user
            well.save()
    return Response(status=200)

@api_view(['GET'])
def staff_request(request, id=0):
    if id == 0:
        user_districts = request.user.staff.districts.all()
        if len(user_districts) != 0:
            staffs = []
            for district in user_districts:
                for item in Staff.objects.filter(districts=district):
                    if (not (item in staffs)) and item.position.priority == request.user.staff.position.priority - 1:
                        staffs.append(item)
            return Response(StaffSerializer(staffs, many=True).data, status=200)
        else:
            return Response(StaffSerializer(Staff.objects.all(), many=True).data, status=200)
    else:
        staff = Staff.objects.get(pk=id)
        user = staff.user
        response = {"staff": StaffSerializer(staff).data, "user": {'id': user.id, 'username': user.username}}
        return Response(response, status=200)

@api_view(['POST', 'PUT', "DELETE"])
@transaction.atomic
def user_request(request):
    if request.method == 'POST':
        user = UserSerializer(data=request.data.get('user'))
        staff = StaffSerializer(data=request.data.get('staff'))
        is_user_valid = user.is_valid()
        is_staff_valid = staff.is_valid()
        if is_user_valid and is_staff_valid:
            # add to group by position
            position = Position.objects.get(pk=request.data.get('staff').get('position'))

            user_data = request.data.get('user')
            user_instance = User.objects.create_user(username=user_data.get('username'), password=user_data.get('password'))
            user_instance.groups.add(position.group.id)

            staff.save()
            staff.instance.user = user_instance
            staff.save()

            response = {"user": user_instance.id, "staff": staff.instance.id}
            return Response(response, status=200)
        else:
            error_msg = {}
            if not is_user_valid:
                error_msg['user_errors'] = user.errors
            if not is_staff_valid:
                error_msg['staff_errors'] = staff.errors
            return Response(error_msg, status=406)
    try:
        staff_instance = Staff.objects.get(pk=request.data.get('staff').get('id'))
        user_instance = User.objects.get(pk=request.data.get('user').get('id'))
    except:
        return Response("Not Found", status=404)

    if request.method == 'PUT':
        user = UserSerializer(data=request.data.get('user'), instance=user_instance)
        staff = StaffSerializer(data=request.data.get('staff'), instance=staff_instance)

        is_user_valid = user.is_valid()
        is_staff_valid = staff.is_valid()
        if is_user_valid and is_staff_valid:
            # add to group by position
            new_position = Position.objects.get(pk=request.data.get('staff').get('position'))

            user_instance.groups.clear()
            user_instance.groups.add(new_position.group.id)

            staff.save()

            return Response("Ok", status=200)
        else:
            error_msg = {}
            if not is_user_valid:
                error_msg['user_errors'] = user.errors
            if not is_staff_valid:
                error_msg['staff_errors'] = staff.errors
            return Response(error_msg, status=406)

    else:
        user_instance.delete()
        staff_instance.delete()

        return Response("Deleted", status=200)

@api_view(['GET'])
def position_request(reqeust):
    if reqeust.method == 'GET':
        return Response(PositionSerializer(Position.objects.filter(priority__lt = reqeust.user.staff.position.priority), many=True).data, status=200)

@api_view(['GET'])
def group_request(reqeust):
    if reqeust.method == 'GET':
        return Response(GroupSerializer(Group.objects.all(), many=True).data, status=200)

@api_view(['GET'])
def get_user_information(request):
    permissions = get_user_permissions(request.user)
    staff = request.user.staff
    title = f"{staff.last_name} {staff.first_name[0]}. {staff.first_name[0] + '.' if len(staff.first_name) > 0 else ''}"
    return Response({'permissions': permissions, 'user_title': title}, status=200)

@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def ugv_request(request, well_id=0):
    try:
        if request.method == 'GET':
            if well_id != 0:
                well = Well.objects.get(pk = well_id)
                ugvs = Ugv.objects.filter(well=well)
                return Response(UgvSerializer(ugvs, many=True).data, status=200)
            
            # send all ugvs which the user has
            ugvs = Ugv.objects.filter(staff=request.user.staff)
            return Response(UgvSerializer(ugvs, many=True).data, status=200)
        # Create object
        if request.method == 'POST':
            request.data['staff']=request.user.staff.id
            serializer = UgvSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)
        
        # find suitable instance
        try:
            instance = Ugv.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            request.data['staff']=request.user.staff.id 
            serializer = UgvSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def mgv_request(request, well_id=0):
    try:
        if request.method == 'GET':
            if well_id != 0:
                well = Well.objects.get(pk = well_id)
                mgvs = Mgv.objects.filter(well=well)
                return Response(MgvSerializer(mgvs, many=True).data, status=200)
            
            # send all ugvs which the user has
            mgvs = Mgv.objects.filter(staff=request.user.staff)
            return Response(UgvSerializer(mgvs, many=True).data, status=200)
        # Create object
        if request.method == 'POST':
            request.data['staff']=request.user.staff.id
            serializer = MgvSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)
        
        # find suitable instance
        try:
            instance = Mgv.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            request.data['staff']=request.user.staff.id 
            serializer = MgvSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)


@authentication_classes([BasicAuthentication])
@permission_classes([IsAuthenticated])
@api_view(['POST'])
def ugv_from_weighter(request):
    print(request.data)
    print(request)
    return Response("Ok", status=200)


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000

    def get_paginated_response(self, data):
        return Response({
            'links': {
               'next': self.get_next_link(),
               'previous': self.get_previous_link()
            },
            'count': self.page.paginator.count,
            'pagesCount': self.page.paginator.num_pages,
            'results': data
        })

class MustorSoilView(ListAPIView):
    #queryset = MusterSoil.objects.order_by('id').all()
    #queryset = MusterSoil.objects.all()
    serializer_class = MusterSoilSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        district = self.request.query_params.get('district')
        return MusterSoil.objects.filter(well__farm__district=district).order_by('id').all()


@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def muster_soil_request(request, id=0):
    try:
        if request.method == 'GET':
            if id == 0:
                if request.user.has_perm('main.add_analysissoil'):
                    return Response(MusterSoilSerializer(MusterSoil.objects.all(), many=True).data, status=200)
                else:
                    return Response(MusterSoilSerializer(MusterSoil.objects.filter(user=request.user), many=True).data, status=200)
            else:
                try:
                    obj = MusterSoil.objects.get(pk=id)
                    return Response(MusterSoilSerializer(obj).data, status=200)
                except:
                    return Response("Not Found", status=404)                    
        # Create object
        if request.method == 'POST':
            request.data['user']=request.user.id
            serializer = MusterSoilSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                analysis_soil_analysis = AnalysisSoil()
                analysis_soil_analysis.muster_soil = serializer.instance
                analysis_soil_analysis.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)
        
        # find suitable instance
        try:
            instance = MusterSoil.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            request.data['user']=request.user.id 
            serializer = MusterSoilSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(['GET'])
def farms_by_user_district(request):
    try:
        query = []
        for district in request.user.staff.districts.all():
            for farm in Farm.objects.filter(district=district):
                query.append(farm)
        return Response(FarmSerializer(query, many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(['GET'])
def well_by_farm(request, farm_id):
    try:
        return Response(WellSerializer(Well.objects.filter(farm=farm_id), many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(["GET"])
def salt_degree_request(request):
    try:
        return Response(SaltDegreeSerializer(SaltDegree.objects.all(), many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(["GET", "POST", "PUT", "DELETE"])
def soil_deep_request(request, mustersoilid=0):
    try:
        if request.method == "GET":
            if mustersoilid != 0:
                return Response(SoilDeepSerializer(SoilDeep.objects.filter(muster_soil=mustersoilid), many=True).data, status=200)
            else:
               return Response(status=400)
        
        # Create object
        if request.method == 'POST':
            serializer = SoilDeepSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        try:
            instance = SoilDeep.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        # Update object
        if request.method == 'PUT':
            serializer = SoilDeepSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(["GET"])
def soil_type_request(request):
    try:
        return Response(SoilTypeSerializer(SoilType.objects.all(), many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(["GET"])
def crop_type_request(request):
    try:
        return Response(CropTypeSerializer(CropType.objects.all(), many=True).data, status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(["GET", "POST", "PUT", "DELETE"])
def analysis_soil_request(request, mustersoilid=0):
    try:
        if request.method == "GET":
            if mustersoilid != 0:
                return Response(AnalysisSoilSerializer(AnalysisSoil.objects.filter(muster_soil=mustersoilid), many=True).data, status=200)
            else:
                Response(status=400)
        
        # Create object
        if request.method == 'POST':
            serializer = AnalysisSoilSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        try:
            instance = AnalysisSoil.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)
        
        
        # Update object
        if request.method == 'PUT':
            request.data['user']=request.user.id

            serializer = AnalysisSoilSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        Response(status=500)

def get_degrees_by_decades(well, year, month):
    ugvs = Ugv.objects.filter(well=well, date__year=year, date__month=month)

    first_dec = 0
    first_dec_count = 0

    second_dec = 0
    second_dec_count = 0

    third_dec = 0
    third_dec_count = 0

    for ugv in ugvs:
        if ugv.date.day <= 10:
            first_dec += ugv.degree
            first_dec_count += 1
        elif ugv.date.day <= 20:
            second_dec += ugv.degree
            second_dec_count += 1
        else:
            third_dec += ugv.degree
            third_dec_count += 1

        if first_dec_count != 0:
            first_dec /= first_dec_count
        if second_dec_count != 0:
            second_dec /= second_dec_count
        if third_dec_count != 0:
            third_dec /= third_dec_count

    return first_dec, second_dec, third_dec

def notZero(val):
    return val != 0 and val is not None
def notZeroMean(*vals):
    not_zero_count = 0
    not_zero_sum = 0
    for val in vals:
        if notZero(val):
            not_zero_count += 1
            not_zero_sum += val
    if not_zero_count != 0:
        return not_zero_sum / not_zero_count
    else:
        return 0

def get_closest_to_date(Class, date, well):
    greater = Class.objects.filter(well=well, date__gte=date).order_by("date").first()
    less = Class.objects.filter(well=well, date__lte=date).order_by("-date").first()

    if greater and less:
        return greater if abs(greater.date - date) < abs(less.date - date) else less
    elif greater is not None:
        return greater
    elif less is not None:
        return less
    else:
        return None

@api_view(['GET'])
def report_6a(request, district, date):

    months_dict = {
        1: 'январь',
        2: 'февраль',
        3: 'март',
        4: 'апрель',
        5: 'май',
        6: 'июнь',
        7: 'июль',
        8: 'августь',
        9: 'сентябрь',
        10: 'откябрь',
        11: 'ноябрь',
        12: 'декабрь',
    }

    year, month, _ = date.split('-')
    year = int(year)
    month = int(month)
    wb = load_workbook(os.path.join(
        settings.BASE_DIR, 'reports', 'a6.xlsx'))
    sheet = wb.active

    center_align = Alignment(horizontal='center', vertical='center')
    vertical_text = Alignment(
        horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    general_style = NamedStyle(
        name='a_style', alignment=center_align, border=border)
    general_bold_style = NamedStyle(
        name='general_bold_style', alignment=center_align, font=Font(b=True), border=border)
    bc_style = NamedStyle(
        name='bc_style', alignment=vertical_text, border=border)
    center_no_border = NamedStyle(
        name='center_no_border', alignment=center_align)

    obj_dist = District.objects.get(pk=district)
    sheet.title = obj_dist.name
    sheet['A1'] = f'{obj_dist.name} туманидаги мелиоратив кудуклардаги ер ости сувларини жойлашиши хакида '
    sheet['A2'] = f'МАЪЛУМОТ {months_dict[month]} ойи {year} й'
    sheet['J4'] = f'{year - 1} й {months_dict[month]} \n уртача чукурлиги см'
    sheet.merge_cells('A1:L1')
    sheet.merge_cells('A2:L2')
    sheet.merge_cells('J4:J5')
    sheet['A1'].style = center_no_border
    sheet['A2'].style = center_no_border
    sheet['J4'].style = bc_style

    farms = Farm.objects.filter(district=district)
    farm_number = 1
    row_number = 7
    try:
        for farm in farms:

            wells = Well.objects.filter(farm=farm)
            well_count = len(wells)

            if well_count > 0:

                sheet[f'A{row_number}'] = farm_number
                sheet[f'B{row_number}'] = farm.name

                sheet[f'A{row_number}'].style = general_style
                sheet[f'B{row_number}'].style = bc_style
                sheet[f'C{row_number}'].style = bc_style

                if wells[0].user is not None:
                    staff = wells[0].user.staff
                    sheet[f'C{row_number}'] = staff.last_name + \
                        " " + staff.first_name
                else:
                    sheet[f'C{row_number}'] = "Staff not found"

                sheet.merge_cells(
                    f'A{row_number}:A{row_number + well_count + 1}')
                sheet.merge_cells(
                    f'B{row_number}:B{row_number + well_count - 1}')
                sheet.merge_cells(
                    f'C{row_number}:C{row_number + well_count - 1}')
                sheet.merge_cells(
                    f'B{row_number + well_count}:C{row_number + well_count}')
                sheet.merge_cells(
                    f'B{row_number + well_count + 1}:C{row_number + well_count + 1}')

                sheet[f'B{row_number + well_count}'] = 'Уртача'
                sheet[f'B{row_number + well_count + 1}'] = 'Жами'

                sum_area = 0
                sum_first_dec = 0
                sum_second_dec = 0
                sum_third_dec = 0
                sum_average_dec = 0
                sum_last_year = 0

                # actual well count for each situation
                well_count_first_dec = 0
                well_count_second_dec = 0
                well_count_third_dec = 0
                well_count_mean_dec = 0
                well_count_last_year = 0

                for well in wells:
                    sum_area += well.area
                    if notZero(well.area):
                        sheet[f'D{row_number}'] = well.area
                    if notZero(well.number):
                        sheet[f'E{row_number}'] = well.number
                    if notZero(well.label):
                        sheet[f'L{row_number}'] = well.label

                    first_dec, second_dec, third_dec = get_degrees_by_decades(
                        well, year, month)

                    sum_first_dec += first_dec
                    sum_second_dec += second_dec
                    sum_third_dec += third_dec
                    average = notZeroMean(first_dec, second_dec, third_dec)
                    sum_average_dec += average

                    if notZero(first_dec):
                        well_count_first_dec += 1
                        sheet[f'F{row_number}'] = f'{first_dec:.2f}'
                    if notZero(second_dec):
                        well_count_second_dec += 1
                        sheet[f'G{row_number}'] = f'{second_dec:.2f}'
                    if notZero(third_dec):
                        well_count_third_dec += 1
                        sheet[f'H{row_number}'] = f'{third_dec:.2f}'
                    if notZero(average):
                        well_count_mean_dec += 1
                        sheet[f'I{row_number}'] = f'{average:.2f}'

                    last_year_degree = notZeroMean(
                        *get_degrees_by_decades(well, year, month))
                    sum_last_year += last_year_degree
                    if notZero(last_year_degree):
                        well_count_last_year += 1
                        sheet[f'J{row_number}'] = f'{last_year_degree:.2f}'

                    closest_ugv = get_closest_to_date(
                        MusterPumping, datetime.date(year - 1, month, 1), well)
                    if closest_ugv is not None:
                        sheet[f'K{row_number}'] = closest_ugv.bottom
                    row_number += 1

                sheet[f'F{row_number + 1}'] = f'{sum_first_dec:.2f}'
                sheet[f'G{row_number + 1}'] = f'{sum_second_dec:.2f}'
                sheet[f'H{row_number + 1}'] = f'{sum_third_dec:.2f}'
                sheet[f'I{row_number + 1}'] = f'{sum_average_dec:.2f}'
                sheet[f'J{row_number + 1}'] = f'{sum_last_year:.2f}'

                if sum_first_dec != 0:
                    sum_first_dec /= well_count_first_dec
                if sum_second_dec != 0:
                    sum_second_dec /= well_count_second_dec
                if sum_third_dec != 0:
                    sum_third_dec /= well_count_third_dec
                if sum_average_dec != 0:
                    sum_average_dec /= well_count_mean_dec
                if sum_last_year != 0:
                    sum_last_year /= well_count_last_year

                sheet[f'F{row_number}'] = f'{sum_first_dec:.2f}'
                sheet[f'G{row_number}'] = f'{sum_second_dec:.2f}'
                sheet[f'H{row_number}'] = f'{sum_third_dec:.2f}'
                sheet[f'I{row_number}'] = f'{sum_average_dec:.2f}'
                sheet[f'J{row_number}'] = f'{sum_last_year:.2f}'
                sheet[f'D{row_number + 1}'] = f'{sum_area:.2f}'
                sheet[f'E{row_number + 1}'] = well_count

                # General styling
                for i in range(row_number - well_count, row_number):
                    for letter in ('D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'):
                        sheet[f'{letter}{i}'].style = general_style

                # styles for ever well
                for letter in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'):
                    sheet[f'{letter}{row_number}'].style = general_bold_style
                    sheet[f'{letter}{row_number + 1}'].style = general_bold_style
                farm_number += 1
                row_number += 2
    except:
        traceback.print_exc()
        raise
    #wb.save('D:/a6_res.xlsx')
    #wb.close()

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
    return response

@api_view(['GET'])
def report_soil_analysis(request, district, date):
    months_dict = {
        1: 'январь',
        2: 'февраль',
        3: 'март',
        4: 'апрель',
        5: 'май',
        6: 'июнь',
        7: 'июль',
        8: 'августь',
        9: 'сентябрь',
        10: 'откябрь',
        11: 'ноябрь',
        12: 'декабрь',
    }

    year, month, _ = date.split('-')
    year = int(year)
    month = int(month)
    wb = load_workbook(os.path.join(
        settings.BASE_DIR, 'reports', 'report-soil-analysis.xlsx'))
    sheet = wb.active

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    vertical_text = Alignment(
        horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    general_style = NamedStyle(
        name='a_style', alignment=center_align, border=border)
    
    bc_style = NamedStyle(
        name='bc_style', alignment=vertical_text, border=border)
    center_no_border = NamedStyle(
        name='center_no_border', alignment=center_align)

    obj_dist = District.objects.get(pk=district)
    sheet.title = obj_dist.name
    sheet['A1'] = f'Отчет'
    sheet['A2'] = f'Результатов хим-анализов почвагрунтов по солевому опробованию за {months_dict[month]} месяц {year} года из хозяйств {obj_dist.name} района'
    sheet.merge_cells('A1:K1')
    sheet.merge_cells('A2:K2')
    sheet['A1'].style = center_no_border
    sheet['A2'].style = center_no_border

    farms = Farm.objects.filter(district=district)
    farm_number = 1
    row_number = 9
    try:
        for farm in farms:

            wells = Well.objects.filter(farm=farm)
            merging_row = row_number

            print('row_number', row_number)

            sheet[f'A{row_number}'] = farm_number
            sheet[f'B{row_number}'] = farm.name

            sheet[f'A{row_number}'].style = general_style
            sheet[f'B{row_number}'].style = bc_style

                
            for well in wells:
                sheet[f'C{row_number}'] = well.number
                mss = MusterSoil.objects.filter(well=well)

                for ms in mss:
                    sheet[f'D{row_number}'] = ms.pit_no
                    sheet[f'N{row_number}'] = ms.pit_no

                    if hasattr(ms, 'analysissoil'):
                        analysis = ms.analysissoil
                        sheet[f'E{row_number}'] = analysis.electric_wire
                        sheet[f'O{row_number}'] = analysis.electric_wire

                        # if the chemical analysis was done
                        if analysis.hco3 != 0:
                            sheet[f'F{row_number}'] = analysis.hco3 * 0.012
                            sheet[f'G{row_number}'] = analysis.cl * 0.099
                            sheet[f'H{row_number}'] = analysis.so4 * 0.82
                            sheet[f'I{row_number}'] = analysis.ca * 0.05
                            sheet[f'J{row_number}'] = analysis.mg * 0.03

                            sheet[f'P{row_number}'] = analysis.hco3
                            sheet[f'Q{row_number}'] = analysis.cl
                            sheet[f'R{row_number}'] = analysis.so4
                            sheet[f'S{row_number}'] = analysis.ca
                            sheet[f'T{row_number}'] = analysis.mg

                            sheet[f'F{row_number + 1}'] = analysis.hco3 * 0.197
                            sheet[f'G{row_number + 1}'] = analysis.cl * 2.792
                            sheet[f'H{row_number + 1}'] = analysis.so4 * 0.82
                            sheet[f'I{row_number + 1}'] = analysis.ca * 2.494
                            sheet[f'J{row_number + 1}'] = analysis.mg * 2.467

                            sheet[f'J{row_number + 1}'] = analysis.hco3 * 0.197 + analysis.cl * 2.792 + analysis.so4 * 0.82 - analysis.ca * 2.494 - analysis.mg * 2.467

                            sheet[f'J{row_number}'] = 0.023 * (analysis.hco3 * 0.197 + analysis.cl * 2.792 + analysis.so4 * 0.82 - analysis.ca * 2.494 - analysis.mg * 2.467)

                            row_number += 1
                    row_number += 1

                row_number += 1

                # General styling
                for i in range(merging_row, row_number):
                    for letter in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
                        sheet[f'{letter}{i}'].style = general_style
                    
                for i in range(merging_row, row_number):
                    for letter in ('N', 'O', 'P', 'Q', 'R', 'S', 'T'):
                        sheet[f'{letter}{i}'].style = general_style

            farm_number += 1
            sheet.merge_cells(f'A{merging_row}:A{row_number - 1}')
            sheet.merge_cells(f'B{merging_row}:B{row_number - 1}')

    except:
        traceback.print_exc()
        #raise
    #wb.save('D:/a6_res.xlsx')
    #wb.close()

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
    return response

@api_view(['POST'])
def get_wells(request):
    return Response(WellSerializer(Well.objects.filter(pk__in=request.data['ids']), many=True).data, status=200)

@api_view(['POST'])
def get_farms(request):
    return Response(FarmSerializer(Farm.objects.filter(pk__in=request.data['ids']), many=True).data, status=200)


@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def yield_request(request, id=0):
    try:
        if request.method == 'GET':
            if id != 0:
                ys = YieldSize.objects.get(pk=id)
                return Response(YieldSizeSerializer(ys).data, status=200)
            else:
                return Response(YieldSizeSerializer(YieldSize.objects.order_by('id').all(), many=True).data, status=200)
        # Create object
        if request.method == 'POST':
            serializer = YieldSizeSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        print(request.data)
        try:
            instance = YieldSize.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)

        # Update object
        if request.method == 'PUT':
            serializer = YieldSizeSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)


@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def water_request(request, id=0):
    try:
        if request.method == 'GET':
            if id != 0:
                ys = WaterSize.objects.get(pk=id)
                return Response(WaterSize(ys).data, status=200)
            else:
                return Response(WaterSizeSerializer(WaterSize.objects.all(), many=True).data, status=200)
        # Create object
        if request.method == 'POST':
            serializer = WaterSizeSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        try:
            instance = WaterSize.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)

        # Update object
        if request.method == 'PUT':
            serializer = WaterSizeSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)


@api_view(['POST', 'DELETE', 'PUT', 'GET'])
def water_norm_request(request, id=0):
    try:
        if request.method == 'GET':
            if id != 0:
                ys = WaterNorm.objects.get(pk=id)
                return Response(WaterNorm(ys).data, status=200)
            else:
                return Response(WaterNormSerializer(WaterNorm.objects.all(), many=True).data, status=200)
        # Create object
        if request.method == 'POST':
            serializer = WaterNormSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data.get('id'), status=201)
            else:
                return Response(serializer.errors, status=406)

        # find suitable instance
        try:
            instance = WaterNorm.objects.get(id=request.data.get('id'))
        except:
            return Response(status=404)

        # Update object
        if request.method == 'PUT':
            serializer = WaterNormSerializer(instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(status=200)
            else:
                return Response(serializer.errors, status=406)
        else:
            instance.delete()
            return Response('Deleted', status=200)
    except:
        traceback.print_exc()
        return Response(status=500)

@api_view(['GET'])
def well_tool_data(request, imei):
    last_tool_data = WellToolData.objects.filter(imei=imei).last()
    if last_tool_data:
        return Response(WellToolDataSerializer(last_tool_data).data, status=200)
    else:
        return Response('Not found', status=200)


#@api_view(['GET'])
def json_query(request):
    kargs = {}
    for key in request.GET:
        vals = request.GET.get(key)
        if len(vals) > 0:
            kargs[key] = vals[0]
    #return Response(WellSerializer(Well.objects.filter(**kargs), many=True), status=200)
    return HttpResponse(str(kargs), status=200)
